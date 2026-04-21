# -*- coding: utf-8 -*-
"""Importa planilhas .xlsx antigas (geradas pelo write_excel) para o results_store.

Para cada arquivo:
- Gera run_id a partir do timestamp do nome.
- Extrai canais da aba 'Canais (Resumo)' e vídeos da aba 'Vídeos' (já com channel_id).
- Pega URLs reais dos hyperlinks.
- Monta o mesmo payload que _finalize_scored_run gera.
- Grava em resultados_buscas.json (sem duplicar se run_id já existe).
"""
import re
from datetime import datetime
from pathlib import Path

from . import config
from . import results_store


FILENAME_RE = re.compile(r"relatorio_canais_(\d{8})_(\d{4,6})\.xlsx$", re.IGNORECASE)


def _parse_run_id_from_name(path: Path) -> tuple[str, str]:
    """Devolve (run_id, created_at_iso) inferidos do nome do arquivo, ou do mtime se falhar."""
    m = FILENAME_RE.search(path.name)
    if m:
        date_part, time_part = m.group(1), m.group(2)
        # Normaliza time_part para HHMMSS
        if len(time_part) == 4:
            time_part += "00"
        run_id = f"{date_part}_{time_part}"
        try:
            dt = datetime.strptime(run_id, "%Y%m%d_%H%M%S")
            return run_id, dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    # Fallback: mtime do arquivo
    mt = datetime.fromtimestamp(path.stat().st_mtime)
    return mt.strftime("%Y%m%d_%H%M%S"), mt.strftime("%Y-%m-%d %H:%M:%S")


def _safe_int(x):
    try:
        return int(x)
    except (TypeError, ValueError):
        return None


def _safe_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _cell_url(sh, row: int, col: int):
    """Retorna o target do hyperlink na célula (1-based), ou None."""
    try:
        c = sh.cell(row=row, column=col)
    except Exception:
        return None
    if c.hyperlink is not None:
        return c.hyperlink.target
    v = c.value
    if isinstance(v, str) and v.startswith(("http://", "https://")):
        return v
    return None


def _video_id_from_url(url: str):
    if not url:
        return None
    if "watch?v=" in url:
        return url.split("watch?v=", 1)[1].split("&", 1)[0]
    if "youtu.be/" in url:
        return url.split("youtu.be/", 1)[1].split("?", 1)[0].split("/", 1)[0]
    return None


def _iter_rows(ws):
    """Lê todas as linhas (values + hyperlinks). Retorna lista de linhas, cada uma com
    ambos os valores (texto) e hyperlinks (dict col→url)."""
    from openpyxl.utils.cell import get_column_letter  # noqa: F401
    rows = []
    for row in ws.iter_rows():
        values = [c.value for c in row]
        links = {i + 1: c.hyperlink.target for i, c in enumerate(row) if c.hyperlink is not None}
        rows.append((values, links))
    return rows


def parse_xlsx(path: Path) -> dict | None:
    """Parse um .xlsx Scored para o formato do results_store. Retorna None se não for reconhecível."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl")

    try:
        wb = load_workbook(str(path), read_only=False, data_only=True)
    except Exception as ex:
        raise RuntimeError(f"Falha ao abrir {path.name}: {ex}")

    sheets = wb.sheetnames
    if "Canais (Resumo)" not in sheets:
        return None

    run_id, created_at = _parse_run_id_from_name(path)

    # ----- Aba 'Vídeos' (aba 2): mapa channel_title → channel_id, e lista de vídeos -----
    videos_by_channel = {}  # channel_title → [video_dict, ...]
    channel_id_by_title = {}  # usa o primeiro que aparece para cada título
    if "Vídeos" in sheets:
        ws_v = wb["Vídeos"]
        rows_v = _iter_rows(ws_v)
        # header: Canal, Canal ID, Vídeo, URL, Publicado, Duração, Views, Likes, Coment., Score
        for vals, links in rows_v[1:]:
            if not vals or not vals[0]:
                continue
            ch_title = vals[0]
            ch_id = vals[1] if len(vals) > 1 else None
            if ch_title and ch_id and ch_title not in channel_id_by_title:
                channel_id_by_title[ch_title] = ch_id

            video_title = vals[2] if len(vals) > 2 else None
            video_url = links.get(4)  # coluna 4 = URL
            video_id = _video_id_from_url(video_url)
            pub = vals[4] if len(vals) > 4 else None
            if isinstance(pub, datetime):
                pub = pub.strftime("%Y-%m-%dT00:00:00Z")

            entry = {
                "video_id": video_id,
                "title": video_title,
                "url": video_url,
                "published_at": pub,
                "duration_min": _safe_int(vals[5] if len(vals) > 5 else None),
                "views": _safe_int(vals[6] if len(vals) > 6 else None),
                "likes": _safe_int(vals[7] if len(vals) > 7 else None),
                "comments": _safe_int(vals[8] if len(vals) > 8 else None),
                "potential_score": _safe_float(vals[9] if len(vals) > 9 else None),
            }
            videos_by_channel.setdefault(ch_title, []).append(entry)

    # ----- Aba 'Canais (Resumo)' -----
    ws_c = wb["Canais (Resumo)"]
    rows_c = _iter_rows(ws_c)
    if not rows_c:
        return None

    header = rows_c[0][0] or []
    # Detecta índice de cada coluna por nome (tolerante a versão antiga/nova)
    def col_of(name):
        for i, h in enumerate(header):
            if h and str(h).strip().lower() == name.strip().lower():
                return i
        return None

    idx = {
        "canal": col_of("Canal"),
        "custom": col_of("Custom URL"),
        "criado": col_of("Criado"),
        "idade": col_of("Idade (dias)"),
        "inscritos": col_of("Inscritos"),
        "views_canal": col_of("Views canal"),
        "views_video": col_of("Views/Video (canal)"),
        "videos_qtd": col_of("Vídeos"),
        "median_views": col_of("Mediana views (últimos)"),
        "pct_long": col_of("% longos"),
        "pct_over": col_of("% ≥ min_views"),
        "janela_uploads": col_of("Janela uploads"),
        "aprov_qtd": col_of("Qtd aprovados"),
        "avg_views_aprov": col_of("Média views aprov."),
        "avg_dur_aprov": col_of("Média duração (min)"),
        "score": col_of("Score"),
        "melhor_video": col_of("Melhor vídeo"),
        "url_video": col_of("URL vídeo"),
        "publicado": col_of("Publicado"),
        "views": col_of("Views"),
        "vpd": col_of("Views/dia"),
        "vps": col_of("Views/inscrito"),
        "like_rate": col_of("Like rate"),
        "comment_rate": col_of("Comment rate"),
        "title_score": col_of("TitleScore"),
        "novelty": col_of("Novidade"),
        "uploads_per_week": col_of("Uploads/sem"),
        "vpd_trend": col_of("Tendência VPD"),
    }

    channels = []
    flat_videos = []
    for vals, links in rows_c[1:]:
        if not vals or not vals[1]:  # coluna Canal vazia
            continue

        def g(key):
            i = idx.get(key)
            return vals[i] if (i is not None and i < len(vals)) else None

        ch_title = g("canal")
        ch_id = channel_id_by_title.get(ch_title)
        # Tenta pegar URL do hyperlink coluna "URL" (índice 2, 1-based = 3)
        channel_url = _cell_url_from_header(header, rows_c, rows_c.index((vals, links)) + 1,
                                            "URL") or (f"https://www.youtube.com/channel/{ch_id}" if ch_id else None)

        criado = g("criado")
        if isinstance(criado, datetime):
            criado = criado.strftime("%Y-%m-%dT00:00:00Z")

        publicado = g("publicado")
        if isinstance(publicado, datetime):
            publicado = publicado.strftime("%Y-%m-%dT00:00:00Z")

        # Melhor vídeo URL: tenta pegar hyperlink da coluna "URL vídeo"
        url_video = None
        if idx.get("url_video") is not None:
            url_video = links.get(idx["url_video"] + 1)

        best_video = {
            "video_id": _video_id_from_url(url_video),
            "title": g("melhor_video"),
            "url": url_video,
            "published_at": publicado,
            "views": _safe_int(g("views")),
            "metrics": {
                "vpd": _safe_float(g("vpd")),
                "vps": _safe_float(g("vps")),
                "like": _safe_float(g("like_rate")),
                "comm": _safe_float(g("comment_rate")),
                "title": _safe_float(g("title_score")),
                "novelty": _safe_float(g("novelty")),
            },
        }

        consistency = {
            "views_median": _safe_int(g("median_views")),
            "pct_long": _safe_float(g("pct_long")),
            "pct_over": _safe_float(g("pct_over")),
            "uploads_per_week": _safe_float(g("uploads_per_week")),
            "vpd_trend": _safe_float(g("vpd_trend")),
        }
        janela_uploads = g("janela_uploads") or ""
        if " → " in str(janela_uploads):
            lo, hi = str(janela_uploads).split(" → ", 1)
            consistency["date_min"] = lo.strip() if lo.strip() != "—" else None
            consistency["date_max"] = hi.strip() if hi.strip() != "—" else None

        top_videos = []
        for v in videos_by_channel.get(ch_title, []):
            top_videos.append({
                "video_id": v["video_id"],
                "title": v["title"],
                "url": v["url"],
                "published_at": v["published_at"],
                "duration_min": v["duration_min"],
                "views": v["views"],
                "likes": v["likes"],
                "comments": v["comments"],
                "potential_score": v["potential_score"],
            })

        channels.append({
            "channel_id": ch_id,
            "channel_title": ch_title,
            "channel_url": channel_url,
            "custom_url": g("custom"),
            "created_at": criado,
            "age_days": _safe_int(g("idade")),
            "subscribers": _safe_int(g("inscritos")),
            "views_total": _safe_int(g("views_canal")),
            "video_count": _safe_int(g("videos_qtd")),
            "views_per_video": _safe_int(g("views_video")),
            "consistency": consistency,
            "approved_count": _safe_int(g("aprov_qtd")),
            "avg_views_approved": _safe_float(g("avg_views_aprov")),
            "avg_dur_approved": _safe_float(g("avg_dur_aprov")),
            "score": _safe_float(g("score")),
            "best_video": best_video,
            "top_videos": top_videos,
        })

        # Flatten para aba Vídeos do run
        for v in top_videos:
            flat_videos.append({
                "video_id": v["video_id"],
                "channel_id": ch_id,
                "channel_title": ch_title,
                "title": v["title"],
                "url": v["url"],
                "published_at": v["published_at"],
                "duration_min": v["duration_min"],
                "views": v["views"],
                "likes": v["likes"],
                "comments": v["comments"],
                "potential_score": v["potential_score"],
                "vpd": (v["views"] / max(1, _days_since(v["published_at"])))
                       if v.get("published_at") and v.get("views") else None,
            })

    run = {
        "run_id": run_id,
        "created_at": created_at,
        "mode": "scored",
        "params": {"imported_from_xlsx": str(path.name)},
        "terms_used": [],
        "uploads_sample": config.CFG.get("UPLOADS_SAMPLE", 6),
        "quota_used": 0,
        "channels_count": len(channels),
        "videos_count": len(flat_videos),
        "excel_path": str(path),
        "channels": channels,
        "videos": flat_videos,
    }
    return run


def _cell_url_from_header(header, rows, data_row_idx, col_name):
    """Busca URL de hyperlink pela coluna com nome `col_name`. data_row_idx é 1-based na lista rows."""
    for i, h in enumerate(header):
        if h and str(h).strip().lower() == col_name.strip().lower():
            try:
                return rows[data_row_idx][1].get(i + 1)
            except Exception:
                return None
    return None


def _days_since(iso_str):
    if not iso_str:
        return 1
    try:
        if isinstance(iso_str, datetime):
            dt = iso_str
        else:
            # Aceita tanto "2025-10-30" quanto "2025-10-30T00:00:00Z"
            s = str(iso_str).replace("Z", "")
            try:
                dt = datetime.fromisoformat(s)
            except ValueError:
                dt = datetime.strptime(s[:10], "%Y-%m-%d")
        delta = datetime.now() - dt
        return max(1, delta.days)
    except Exception:
        return 1


def import_all_from(folder: Path = None, status_cb=None, skip_existing: bool = True) -> dict:
    """Importa todos os relatorio_canais_*.xlsx de `folder`. Retorna stats dict."""
    folder = folder or config.DATA_DIR
    folder = Path(folder)

    def _log(m):
        if status_cb:
            status_cb(m)

    existing_runs = {r.get("run_id") for r in results_store.load_runs().get("runs", [])}

    files = sorted(folder.glob("relatorio_canais_*.xlsx"))
    if not files:
        _log("Nenhum relatorio_canais_*.xlsx encontrado.")
        return {"imported": 0, "skipped": 0, "errors": 0, "total": 0}

    _log(f"Encontrados {len(files)} arquivo(s).")
    imported = skipped = errors = 0
    err_details = []
    for p in files:
        try:
            run_id_preview, _ = _parse_run_id_from_name(p)
        except Exception:
            run_id_preview = p.name
        if skip_existing and run_id_preview in existing_runs:
            skipped += 1
            _log(f"  ⏭️  {p.name} (já importado)")
            continue
        try:
            run = parse_xlsx(p)
            if not run:
                errors += 1
                err_details.append(f"{p.name}: formato não reconhecido")
                continue
            if run["run_id"] in existing_runs:
                skipped += 1
                _log(f"  ⏭️  {p.name} (run_id já existe)")
                continue
            results_store.append_run(run)
            existing_runs.add(run["run_id"])
            imported += 1
            _log(f"  ✅ {p.name} ({run['channels_count']} canais)")
        except Exception as ex:
            errors += 1
            err_details.append(f"{p.name}: {ex}")
            _log(f"  ❌ {p.name}: {ex}")

    _log(f"Resumo: importados={imported}, pulados={skipped}, erros={errors}.")
    return {
        "imported": imported, "skipped": skipped, "errors": errors,
        "total": len(files), "errors_details": err_details,
    }
